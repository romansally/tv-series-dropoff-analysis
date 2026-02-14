"""Generate Excel QA reconciliation workbook.

Usage:
    python pipeline/03_generate_qa_workbook.py            # Process data/ CSVs
    python pipeline/03_generate_qa_workbook.py --sample   # Process data/sample/ CSVs
"""

import argparse
import math
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from pipeline.config import OUTPUT_DIR, SAMPLE_DIR, DIM_SHOW_PATH

EXCEL_DIR = PROJECT_ROOT / "excel"

# Conditional fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BOLD_FONT = Font(bold=True)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate QA reconciliation workbook"
    )
    parser.add_argument(
        "--sample",
        action="store_true",
        help="Use data/sample/ CSVs as input",
    )
    return parser.parse_args()


def load_and_validate(data_dir: Path, is_sample: bool):
    """Load input CSVs with fail-fast validation. Returns dict of DataFrames."""
    files = {
        "episodes_filtered": data_dir / "episodes_filtered.csv",
        "agg_season_kpis": data_dir / "agg_season_kpis.csv",
        "shark_jump_results": data_dir / "shark_jump_results.csv",
        "dim_show_category": DIM_SHOW_PATH,
    }

    # Check existence and non-empty
    for name, path in files.items():
        if not path.exists():
            print(f"ERROR: Required file not found: {path}")
            sys.exit(1)

    dfs = {name: pd.read_csv(path) for name, path in files.items()}

    for name, df in dfs.items():
        if len(df) == 0:
            print(f"ERROR: {files[name]} has 0 data rows")
            sys.exit(1)

    # Assert required columns
    required_cols = {
        "episodes_filtered": [
            "episode_tconst", "show_tconst", "season_num",
            "episode_num", "avg_rating", "num_votes",
        ],
        "agg_season_kpis": [
            "show_tconst", "season_num", "episode_count", "weighted_rating",
        ],
        "shark_jump_results": ["show_tconst", "shark_jump_season"],
        "dim_show_category": ["show_tconst", "title", "category"],
    }

    for name, cols in required_cols.items():
        missing = set(cols) - set(dfs[name].columns)
        if missing:
            print(f"ERROR: {name} missing required columns: {missing}")
            sys.exit(1)

    return dfs


def auto_width(ws):
    """Approximate auto-width for all columns."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def freeze_and_bold_header(ws):
    """Bold header row and freeze top row."""
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = BOLD_FONT


# ─── Tab 2: EpisodeCountPivot ──────────────────────────────────


def build_episode_count_pivot(ep_df, kpi_df, dim_df):
    """Returns (DataFrame for tab, n_pass, n_fail)."""
    # Count from episodes_filtered
    ep_counts = (
        ep_df.groupby(["show_tconst", "season_num"])
        .size()
        .reset_index(name="ep_count_from_episodes")
    )

    # Count from KPIs
    kpi_counts = kpi_df[["show_tconst", "season_num", "episode_count"]].copy()
    kpi_counts = kpi_counts.rename(columns={"episode_count": "ep_count_from_kpis"})

    # Outer join
    merged = ep_counts.merge(
        kpi_counts, on=["show_tconst", "season_num"], how="outer"
    )

    # Fill NaN for missing sides (they'll cause match=False)
    merged["ep_count_from_episodes"] = merged["ep_count_from_episodes"].fillna(-1).astype(int)
    merged["ep_count_from_kpis"] = merged["ep_count_from_kpis"].fillna(-1).astype(int)

    # Replace sentinel -1 with None for display, but compute diff first
    merged["diff"] = merged["ep_count_from_episodes"] - merged["ep_count_from_kpis"]
    merged["match"] = merged["diff"] == 0

    # Replace -1 sentinels back to None for display
    merged.loc[merged["ep_count_from_episodes"] == -1, "ep_count_from_episodes"] = None
    merged.loc[merged["ep_count_from_kpis"] == -1, "ep_count_from_kpis"] = None

    # Add title
    merged = merged.merge(
        dim_df[["show_tconst", "title"]], on="show_tconst", how="left"
    )

    # Deterministic sort
    merged = merged.sort_values(
        ["show_tconst", "season_num"]
    ).reset_index(drop=True)

    cols = [
        "show_tconst", "title", "season_num",
        "ep_count_from_episodes", "ep_count_from_kpis", "diff", "match",
    ]
    merged = merged[cols]

    n_pass = int(merged["match"].sum())
    n_fail = len(merged) - n_pass
    return merged, n_pass, n_fail


def write_episode_count_pivot(ws, df):
    """Write EpisodeCountPivot tab."""
    headers = list(df.columns)
    ws.append(headers)

    for _, row in df.iterrows():
        ws.append(list(row))

    # Conditional red fill on match=False
    match_col = headers.index("match") + 1
    for row_idx in range(2, len(df) + 2):
        cell = ws.cell(row=row_idx, column=match_col)
        if cell.value is False or cell.value == "False":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = RED_FILL

    freeze_and_bold_header(ws)
    auto_width(ws)


# ─── Tab 3: WeightedRatingCheck ─────────────────────────────────


def build_weighted_rating_check(ep_df, kpi_df, dim_df):
    """Returns (list of row-dicts for the tab, n_pass, n_fail)."""
    shows = sorted(ep_df["show_tconst"].unique())
    rows = []
    n_pass = 0
    n_fail = 0

    for show in shows:
        show_kpis = kpi_df[kpi_df["show_tconst"] == show].copy()
        show_title = dim_df.loc[
            dim_df["show_tconst"] == show, "title"
        ].values
        title = show_title[0] if len(show_title) > 0 else show

        # Pick spot-check season: not season 1 if possible, most episodes, tie-break lowest season_num
        candidates = show_kpis.copy()
        if len(candidates[candidates["season_num"] != 1]) > 0:
            candidates = candidates[candidates["season_num"] != 1]

        candidates = candidates.sort_values(
            ["episode_count", "season_num"], ascending=[False, True]
        )
        chosen_season = int(candidates.iloc[0]["season_num"])
        sql_wr = float(candidates.iloc[0]["weighted_rating"])

        # Get episodes for this show+season
        eps = ep_df[
            (ep_df["show_tconst"] == show) & (ep_df["season_num"] == chosen_season)
        ].copy()
        eps = eps.sort_values(
            ["episode_num", "episode_tconst"]
        ).reset_index(drop=True)

        eps["rating_x_votes"] = eps["avg_rating"] * eps["num_votes"]

        for _, ep in eps.iterrows():
            rows.append({
                "show_tconst": show,
                "title": title,
                "season_num": chosen_season,
                "episode_tconst": ep["episode_tconst"],
                "episode_num": int(ep["episode_num"]),
                "avg_rating": ep["avg_rating"],
                "num_votes": int(ep["num_votes"]),
                "rating_x_votes": ep["rating_x_votes"],
                "is_summary": False,
            })

        sum_rxv = eps["rating_x_votes"].sum()
        sum_votes = int(eps["num_votes"].sum())

        if sum_votes == 0:
            manual_wr = None
            diff = None
            passed = False
            note = "sum_votes=0"
        else:
            manual_wr = sum_rxv / sum_votes
            diff = abs(manual_wr - sql_wr)
            passed = diff <= 0.01
            note = ""

        if passed:
            n_pass += 1
        else:
            n_fail += 1

        rows.append({
            "show_tconst": show,
            "title": title,
            "season_num": chosen_season,
            "episode_tconst": "SUMMARY",
            "episode_num": "",
            "avg_rating": "",
            "num_votes": sum_votes,
            "rating_x_votes": sum_rxv,
            "is_summary": True,
            "sum_rating_x_votes": sum_rxv,
            "sum_votes": sum_votes,
            "manual_weighted_rating": manual_wr if manual_wr is not None else "NA",
            "sql_weighted_rating": sql_wr,
            "diff": diff if diff is not None else "NA",
            "pass": passed,
            "note": note,
        })

    return rows, n_pass, n_fail


def write_weighted_rating_check(ws, rows):
    """Write WeightedRatingCheck tab."""
    headers = [
        "show_tconst", "title", "season_num", "episode_tconst", "episode_num",
        "avg_rating", "num_votes", "rating_x_votes",
        "sum_rating_x_votes", "sum_votes", "manual_weighted_rating",
        "sql_weighted_rating", "diff", "pass", "note",
    ]
    ws.append(headers)

    for r in rows:
        row_data = []
        for h in headers:
            val = r.get(h, "")
            row_data.append(val)
        ws.append(row_data)

        row_idx = ws.max_row
        if r.get("is_summary"):
            # Bold summary rows
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).font = BOLD_FONT
            # Red fill if pass=False
            pass_col = headers.index("pass") + 1
            if r.get("pass") is False:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = RED_FILL

    freeze_and_bold_header(ws)
    auto_width(ws)


# ─── Tab 4: DuplicateCheck ──────────────────────────────────────


def build_duplicate_check(ep_df):
    """Returns (DataFrame of duplicates, total_dup_count)."""
    counts = ep_df["episode_tconst"].value_counts()
    dups = counts[counts > 1].reset_index()
    dups.columns = ["episode_tconst", "count"]
    dups = dups.sort_values("episode_tconst").reset_index(drop=True)
    return dups, len(dups)


def write_duplicate_check(ws, dup_df):
    """Write DuplicateCheck tab."""
    total = len(dup_df)
    ws.append([f"Total duplicates: {total}"])
    ws.cell(row=1, column=1).font = BOLD_FONT

    headers = ["episode_tconst", "count"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = BOLD_FONT

    for _, row in dup_df.iterrows():
        ws.append([row["episode_tconst"], int(row["count"])])

    freeze_and_bold_header(ws)
    auto_width(ws)


# ─── Tab 5: SharkJumpSanity ─────────────────────────────────────


def build_shark_jump_sanity(shark_df, kpi_df, dim_df):
    """Returns (DataFrame, n_pass, n_fail)."""
    # Total seasons per show
    total_seasons = (
        kpi_df.groupby("show_tconst")["season_num"]
        .max()
        .reset_index()
        .rename(columns={"season_num": "total_seasons"})
    )

    merged = shark_df.merge(
        dim_df[["show_tconst", "title"]], on="show_tconst", how="left"
    )
    merged = merged.merge(total_seasons, on="show_tconst", how="left")

    merged["flag_suspicious"] = merged["shark_jump_season"].apply(
        lambda x: True if pd.notna(x) and int(x) in (1, 2) else False
    )

    cols = [
        "show_tconst", "title", "shark_jump_season",
        "total_seasons", "flag_suspicious",
    ]
    merged = merged[cols].sort_values("show_tconst").reset_index(drop=True)

    n_fail = int(merged["flag_suspicious"].sum())
    n_pass = len(merged) - n_fail
    return merged, n_pass, n_fail


def write_shark_jump_sanity(ws, df):
    """Write SharkJumpSanity tab."""
    headers = list(df.columns)
    ws.append(headers)

    for _, row in df.iterrows():
        ws.append(list(row))

    # Red fill on flag_suspicious=True
    flag_col = headers.index("flag_suspicious") + 1
    for row_idx in range(2, len(df) + 2):
        cell = ws.cell(row=row_idx, column=flag_col)
        if cell.value is True or cell.value == "True":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = RED_FILL

    freeze_and_bold_header(ws)
    auto_width(ws)


# ─── Tab 6: VoteCountCheck ──────────────────────────────────────


def build_vote_count_check(ep_df, dim_df, is_sample: bool):
    """Returns (DataFrame of 3 episodes, mode_label)."""
    # Sort deterministically
    sorted_eps = ep_df.sort_values(
        ["num_votes", "episode_tconst"], ascending=[True, True]
    ).reset_index(drop=True)

    n = len(sorted_eps)

    # Lowest: first row (ties broken by lowest episode_tconst via sort)
    lowest = sorted_eps.iloc[0]

    # Highest: last row; but ties broken by lowest episode_tconst
    max_votes = sorted_eps["num_votes"].max()
    highest_candidates = sorted_eps[sorted_eps["num_votes"] == max_votes]
    highest = highest_candidates.iloc[0]  # already sorted by episode_tconst ASC

    # Median: floor(n/2) index
    median_idx = n // 2
    median = sorted_eps.iloc[median_idx]

    selected = pd.DataFrame([lowest, median, highest])

    # Add title (show name) from dim
    selected = selected.merge(
        dim_df[["show_tconst", "title"]], on="show_tconst", how="left"
    )

    selected["imdb_url"] = (
        "https://www.imdb.com/title/" + selected["episode_tconst"] + "/"
    )
    selected["imdb_web_num_votes"] = ""
    selected["notes"] = ""

    cols = [
        "episode_tconst", "title", "season_num", "episode_num",
        "pipeline_num_votes", "imdb_url", "imdb_web_num_votes", "notes",
    ]
    selected = selected.rename(columns={"num_votes": "pipeline_num_votes"})

    # Deterministic sort: pipeline_num_votes ASC, then episode_tconst ASC
    selected = selected.sort_values(
        ["pipeline_num_votes", "episode_tconst"], ascending=[True, True]
    ).reset_index(drop=True)

    selected = selected[cols]

    mode_label = "SKIP" if is_sample else "MANUAL"
    return selected, mode_label


def write_vote_count_check(ws, df, is_sample: bool):
    """Write VoteCountCheck tab."""
    start_row = 1
    if is_sample:
        ws.cell(row=1, column=1, value="Synthetic data — skip manual verification.")
        ws.cell(row=1, column=1).font = BOLD_FONT
        start_row = 2

    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=h)
        ws.cell(row=start_row, column=col_idx).font = BOLD_FONT

    for row_offset, (_, row) in enumerate(df.iterrows(), 1):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=start_row + row_offset, column=col_idx, value=row[h])

    ws.freeze_panes = f"A{start_row + 1}"
    auto_width(ws)


# ─── Tab 1: QA_Summary ──────────────────────────────────────────


def write_qa_summary(ws, checks):
    """Write QA_Summary tab. checks is list of (check_name, result, detail)."""
    headers = ["check_name", "result", "detail"]
    ws.append(headers)

    for check_name, result, detail in checks:
        ws.append([check_name, result, detail])

    # Conditional fills
    result_col = 2
    for row_idx in range(2, len(checks) + 2):
        cell = ws.cell(row=row_idx, column=result_col)
        if cell.value == "PASS":
            fill = GREEN_FILL
        elif cell.value == "FAIL":
            fill = RED_FILL
        else:  # MANUAL or SKIP
            fill = YELLOW_FILL
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    freeze_and_bold_header(ws)
    auto_width(ws)


# ─── Main ────────────────────────────────────────────────────────


def main():
    args = parse_args()
    is_sample = args.sample
    data_dir = SAMPLE_DIR if is_sample else OUTPUT_DIR
    mode = "sample" if is_sample else "full"

    print(f"=== QA Workbook Generator (mode: {mode}) ===\n")

    # Load and validate inputs
    # In sample mode, Phase 2 outputs are in OUTPUT_DIR (written by 02_run_sql.py)
    # Episodes are in data_dir (sample dir); KPIs/shark are in OUTPUT_DIR
    dfs = load_and_validate(OUTPUT_DIR, is_sample)
    ep_df = dfs["episodes_filtered"]
    kpi_df = dfs["agg_season_kpis"]
    shark_df = dfs["shark_jump_results"]
    dim_df = dfs["dim_show_category"]

    # But episodes_filtered should come from the same place 02_run_sql reads from
    # For sample mode, 02_run_sql.py reads from SAMPLE_DIR but writes KPIs to OUTPUT_DIR
    # The episodes_filtered.csv in OUTPUT_DIR is written by 01_subset_imdb.py
    # Both modes: all Phase 1+2 outputs are in OUTPUT_DIR

    wb = Workbook()

    # ── Tab 2: EpisodeCountPivot ──────────────────────────────
    ecp_df, ecp_pass, ecp_fail = build_episode_count_pivot(ep_df, kpi_df, dim_df)
    ws2 = wb.create_sheet("EpisodeCountPivot")
    write_episode_count_pivot(ws2, ecp_df)

    # ── Tab 3: WeightedRatingCheck ────────────────────────────
    wr_rows, wr_pass, wr_fail = build_weighted_rating_check(ep_df, kpi_df, dim_df)
    ws3 = wb.create_sheet("WeightedRatingCheck")
    write_weighted_rating_check(ws3, wr_rows)

    # ── Tab 4: DuplicateCheck ─────────────────────────────────
    dup_df, dup_count = build_duplicate_check(ep_df)
    ws4 = wb.create_sheet("DuplicateCheck")
    write_duplicate_check(ws4, dup_df)

    # ── Tab 5: SharkJumpSanity ────────────────────────────────
    sjs_df, sjs_pass, sjs_fail = build_shark_jump_sanity(shark_df, kpi_df, dim_df)
    ws5 = wb.create_sheet("SharkJumpSanity")
    write_shark_jump_sanity(ws5, sjs_df)

    # ── Tab 6: VoteCountCheck ─────────────────────────────────
    vc_df, vc_mode = build_vote_count_check(ep_df, dim_df, is_sample)
    ws6 = wb.create_sheet("VoteCountCheck")
    write_vote_count_check(ws6, vc_df, is_sample)

    # ── Tab 1: QA_Summary (must be first sheet) ───────────────
    # Determine results for each check
    ecp_result = "PASS" if ecp_fail == 0 else "FAIL"
    ecp_detail = f"{ecp_pass} matched, {ecp_fail} mismatched"

    wr_result = "PASS" if wr_fail == 0 else "FAIL"
    wr_detail = f"{wr_pass} passed, {wr_fail} failed"

    dup_result = "PASS" if dup_count == 0 else "FAIL"
    dup_detail = f"{dup_count} duplicate episode_tconst values"

    sjs_result = "PASS" if sjs_fail == 0 else "FAIL"
    sjs_detail = f"{sjs_pass} OK, {sjs_fail} suspicious (season 1 or 2)"

    vc_result = vc_mode  # MANUAL or SKIP
    vc_detail = (
        "Synthetic data — skip manual verification"
        if is_sample
        else "3 episodes selected for manual IMDb verification"
    )

    checks = [
        ("EpisodeCountPivot", ecp_result, ecp_detail),
        ("WeightedRatingCheck", wr_result, wr_detail),
        ("DuplicateCheck", dup_result, dup_detail),
        ("SharkJumpSanity", sjs_result, sjs_detail),
        ("VoteCountCheck", vc_result, vc_detail),
    ]

    ws1 = wb.create_sheet("QA_Summary", 0)
    write_qa_summary(ws1, checks)

    # Remove default "Sheet"
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Write workbook ────────────────────────────────────────
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    filename = "qa_reconciliation_sample.xlsx" if is_sample else "qa_reconciliation.xlsx"
    output_path = EXCEL_DIR / filename
    wb.save(str(output_path))

    # ── STDOUT summary ────────────────────────────────────────
    ep_total = sum(1 for r in wr_rows if not r.get("is_summary"))
    wr_summary_count = sum(1 for r in wr_rows if r.get("is_summary"))
    wr_total_rows = len(wr_rows)

    print(f"Tab 1 (QA_Summary): 5 rows, "
          f"{sum(1 for _, r, _ in checks if r == 'PASS')} passed, "
          f"{sum(1 for _, r, _ in checks if r == 'FAIL')} failed")
    print(f"Tab 2 (EpisodeCountPivot): {len(ecp_df)} rows, "
          f"{ecp_pass} passed, {ecp_fail} failed")
    print(f"Tab 3 (WeightedRatingCheck): {wr_total_rows} rows, "
          f"{wr_pass} passed, {wr_fail} failed")
    print(f"Tab 4 (DuplicateCheck): {dup_count} rows, "
          f"{'0 failed' if dup_count == 0 else f'{dup_count} failed'}, "
          f"{'1 passed' if dup_count == 0 else '0 passed'}")
    print(f"Tab 5 (SharkJumpSanity): {len(sjs_df)} rows, "
          f"{sjs_pass} passed, {sjs_fail} failed")
    print(f"Tab 6 (VoteCountCheck): {len(vc_df)} rows, {vc_mode}")
    print(f"\nQA workbook written to: {output_path}")


if __name__ == "__main__":
    main()
